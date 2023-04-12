from openpyxl import load_workbook

import pprint
wb=load_workbook('produceSales.xlsx')
sheet=wb['Sheet']
producedict={'Celery':1.19,'Garlic':3.07,'Lemon':1.27}

for row in range(2,sheet.max_row):
    produceName = sheet.cell(row=row, column=1).value
    if produceName in producedict:
        sheet.cell(row=row,column=2).value=producedict[produceName]
wb.save('updatedProduceSales.xlsx')

