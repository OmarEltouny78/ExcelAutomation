from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.chart import Reference
wb=load_workbook('pivot_table.xlsx')

sheet=wb['Report']

min_col=wb.active.min_column
min_row=wb.active.min_row
max_col=wb.active.max_column
max_row=wb.active.max_row

barchart=BarChart()

data=Reference(sheet,
               min_col=min_col+1,
               min_row=min_row,
               max_col=max_col,
               max_row=max_row)
catergories=Reference(sheet,
               min_col=min_col,
               min_row=min_row+1,
               max_col=min_col,
               max_row=max_row)

barchart.add_data(data=data,titles_from_data=True)
barchart.set_categories(catergories)
barchart.title='Sakes by Gender'

barchart.style=3

sheet.add_chart(barchart, 'B12')

wb.save('barchart.xlsx')