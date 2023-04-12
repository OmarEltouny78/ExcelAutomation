from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.chart import Reference
from openpyxl.utils import get_column_letter
wb=load_workbook('pivot_table.xlsx')

sheet=wb['Report']

min_col=wb.active.min_column
min_row=wb.active.min_row
max_col=wb.active.max_column
max_row=wb.active.max_row

for i in range(min_col+1,max_col+1):
    letter=(get_column_letter(i))
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}6:{letter}7)'
    sheet[f'{letter}{max_row+1}'].style='Currency'



sheet['B8']='=SUM(B6:B7)'

sheet['B8'].style='Currency'

wb.save('report.xlsx')